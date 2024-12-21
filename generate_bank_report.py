import os
import tkinter as tk
from tkinter import filedialog, messagebox
import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill, Alignment

# Function to read the Excel file
def read_excel_file(file_path):
    return pd.read_excel(file_path)

# Function to calculate totals per currency and add them to the DataFrame
def calculate_totals_and_insert(df):
    currency_totals = df.groupby('Currency')['Total'].sum().reset_index()

    modified_rows = []

    for currency, total in currency_totals.values:
        currency_rows = df[df['Currency'] == currency]
        modified_rows.append(currency_rows)

        total_row = pd.DataFrame({
            'Total': total,
            'Currency': currency
        }, index=[0])

        modified_rows.append(total_row)

    final_df = pd.concat(modified_rows, ignore_index=True)
    return final_df

# Function to split the data into "Banks" and "Others" based on "EXECUTION"
def split_data_by_execution(df):
    banks_df = df[df['Counter Party'].str.contains('EXECUTION', case=False, na=False)]
    others_df = df[~df['Counter Party'].str.contains('EXECUTION', case=False, na=False)]
    return banks_df, others_df

# Function to calculate the grand total for all currencies
def calculate_grand_total(banks_df, others_df):
    # Concatenate the DataFrames without summing any values yet
    combined_df = pd.concat([banks_df, others_df], ignore_index=True)

    # Exclude rows where "Counter Party" contains "Total" and "Classification" is empty
    filtered_df = combined_df[~(
            combined_df['Counter Party'].str.contains('Total', case=False, na=False) &
            combined_df['Classification'].isna()
    )]

    # Group by 'Currency' and sum 'Total', but only after excluding "Total" rows and empty Classification
    grand_total = filtered_df.groupby('Currency')['Total'].sum().reset_index()

    return grand_total

# Function to save DataFrames to Excel
def save_to_excel(banks_df, others_df, output_file):
    # Exclude "Total" rows and rows without Classification when calculating the totals for "Banks" and "Others"
    banks_valid = banks_df[
        ~banks_df['Counter Party'].str.contains('Total', case=False, na=False) &
        banks_df['Classification'].notna()
    ]
    others_valid = others_df[
        ~others_df['Counter Party'].str.contains('Total', case=False, na=False) &
        others_df['Classification'].notna()
    ]

    # Calculate currency totals for "Banks" and "Others" (excluding "Total" rows and missing Classification)
    banks_totals = banks_valid.groupby('Currency')['Total'].sum().reset_index().rename(columns={'Total': 'Banks Total'})
    others_totals = others_valid.groupby('Currency')['Total'].sum().reset_index().rename(columns={'Total': 'Others Total'})

    # Merge the summaries for "Banks" and "Others"
    summary_df = pd.merge(banks_totals, others_totals, on='Currency', how='outer')

    # Calculate the Grand Total as the sum of "Banks Total" and "Others Total"
    summary_df['Grand Total'] = summary_df['Banks Total'].fillna(0) + summary_df['Others Total'].fillna(0)

    # Rearrange columns for clarity: Currency, Banks Total, Others Total, Grand Total
    summary_df = summary_df[['Currency', 'Banks Total', 'Others Total', 'Grand Total']]

    with pd.ExcelWriter(output_file) as writer:
        # Write the filtered data to 'Banks' and 'Others' sheets with totals rows
        banks_df.to_excel(writer, sheet_name='Banks', index=False)
        others_df.to_excel(writer, sheet_name='Others', index=False)

        # Write the updated summary to the 'Summary' sheet
        summary_df.to_excel(writer, sheet_name='Summary', index=False)

    print(f"Excel file saved successfully to {output_file}")
    style_headers_and_columns(output_file)

    # Automatically open the file after saving
    try:
        os.startfile(output_file)  # Windows
    except AttributeError:
        # For other operating systems like macOS and Linux
        os.system(f'open "{output_file}"' if os.name == 'posix' else f'xdg-open "{output_file}"')

def style_headers_and_columns(output_file):
    workbook = load_workbook(output_file)
    for sheet in workbook.sheetnames:
        worksheet = workbook[sheet]
        for cell in worksheet[1]:
            cell.font = Font(bold=True, color="FFFFFF")
            cell.fill = PatternFill(start_color="4F81BD", end_color="4F81BD", fill_type="solid")
            cell.alignment = Alignment(horizontal="left")
        for row in worksheet.iter_rows():
            for cell in row:
                if cell.value is not None:
                    cell.alignment = Alignment(horizontal="left")
        for column_cells in worksheet.columns:
            column_letter = column_cells[0].column_letter
            worksheet.column_dimensions[column_letter].width = 20
    workbook.save(output_file)

# Main function to orchestrate the process
def process_excel_file(input_file, output_file):
    df = read_excel_file(input_file)
    banks_df, others_df = split_data_by_execution(df)
    banks_df_with_totals = calculate_totals_and_insert(banks_df)
    others_df_with_totals = calculate_totals_and_insert(others_df)
    grand_total = calculate_grand_total(banks_df_with_totals, others_df_with_totals)
    save_to_excel(banks_df_with_totals, others_df_with_totals, output_file)

# UI for file selection
def open_file_ui():
    def is_file_open(file_path):
        """Check if a file is open by trying to open it in append mode."""
        try:
            with open(file_path, 'a'):
                return False  # File is not open
        except IOError:
            return True  # File is open

    def process_files():
        input_file = input_entry.get()
        output_file = output_entry.get()
        if input_file and output_file:
            if is_file_open(output_file):
                status_label.config(text=f"Please close '{output_file}' before proceeding.", fg="red", font=("Helvetica", 12, "bold"))
                return

            try:
                process_excel_file(input_file, output_file)
                status_label.config(text="Processing completed successfully!", fg="green", font=("Helvetica", 12, "bold"))
                ask_to_exit()  # Ask the user if they want to exit after successful processing
            except Exception as e:
                status_label.config(text=f"Error: {e}", fg="red", font=("Helvetica", 12, "bold"))
        else:
            status_label.config(text="Please provide both input and output file paths.", fg="red", font=("Helvetica", 12, "bold"))

    def browse_input():
        file_path = filedialog.askopenfilename(filetypes=[("Excel files", "*.xlsx")])
        input_entry.delete(0, tk.END)
        input_entry.insert(0, file_path)

    def browse_output():
        file_path = filedialog.asksaveasfilename(defaultextension=".xlsx", filetypes=[("Excel files", "*.xlsx")])
        output_entry.delete(0, tk.END)
        output_entry.insert(0, file_path)

    def ask_to_exit():
        """Show a dialog asking the user if they want to exit the program."""
        should_exit = messagebox.askyesno("Exit Program", "Processing completed. Do you want to exit the program?")
        if should_exit:
            root.destroy()  # Close the UI if the user chooses "Yes"

    # Main window
    root = tk.Tk()
    root.title("Report Generator")
    root.geometry("500x250")
    root.configure(bg="#f0f8ff")  # Light blue background

    # Title
    title_label = tk.Label(root, text="Excel Processor Tool", font=("Helvetica", 16, "bold"), bg="#f0f8ff", fg="#4b8bbe")
    title_label.pack(pady=10)

    # Input Frame
    input_frame = tk.Frame(root, bg="#f0f8ff")
    input_frame.pack(pady=5)
    tk.Label(input_frame, text="Input File:", font=("Helvetica", 12), bg="#f0f8ff").grid(row=0, column=0, padx=10, pady=5, sticky="e")
    input_entry = tk.Entry(input_frame, width=40)
    input_entry.grid(row=0, column=1, padx=10, pady=5)
    tk.Button(input_frame, text="Browse", command=browse_input, bg="#4b8bbe", fg="white", relief="groove").grid(row=0, column=2, padx=10, pady=5)

    # Output Frame
    output_frame = tk.Frame(root, bg="#f0f8ff")
    output_frame.pack(pady=5)
    tk.Label(output_frame, text="Output File:", font=("Helvetica", 12), bg="#f0f8ff").grid(row=1, column=0, padx=10, pady=5, sticky="e")
    output_entry = tk.Entry(output_frame, width=40)
    output_entry.grid(row=1, column=1, padx=10, pady=5)
    tk.Button(output_frame, text="Browse", command=browse_output, bg="#4b8bbe", fg="white", relief="groove").grid(row=1, column=2, padx=10, pady=5)

    # Process Button
    tk.Button(root, text="Process", command=process_files, bg="#4b8bbe", fg="white", font=("Helvetica", 12, "bold"), relief="raised").pack(pady=10)

    # Status Label
    status_label = tk.Label(root, text="", bg="#f0f8ff", font=("Helvetica", 10))
    status_label.pack(pady=10)

    root.mainloop()



# Run the UI
open_file_ui()
